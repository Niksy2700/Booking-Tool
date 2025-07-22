from flask import Flask, render_template, request, flash
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime, timedelta

app = Flask(__name__)
app.secret_key = 'your_secret_key'

EXCEL_FILE = 'bookings.xlsx'

# Create Excel file if not exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Room", "Date", "Start Time", "End Time"])
    wb.save(EXCEL_FILE)

def is_conflict(room, date, new_start_str, new_end_str):
    new_start = datetime.strptime(f"{date} {new_start_str}", "%Y-%m-%d %H:%M")
    new_end = datetime.strptime(f"{date} {new_end_str}", "%Y-%m-%d %H:%M")

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_name, existing_room, existing_date, existing_start_str, existing_end_str = row
        if existing_room == room and existing_date == date:
            existing_start = datetime.strptime(f"{existing_date} {existing_start_str}", "%Y-%m-%d %H:%M")
            existing_end = datetime.strptime(f"{existing_date} {existing_end_str}", "%Y-%m-%d %H:%M")
            if new_start < existing_end and new_end > existing_start:
                return True
    return False

def read_bookings():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    return list(ws.iter_rows(min_row=2, values_only=True))

@app.route('/', methods=['GET', 'POST'])
def book_room():
    bookings = read_bookings()

    if request.method == 'POST':
        name = request.form['name']
        room = request.form['room']
        date = request.form['date']
        start_time = request.form['start_time']
        end_time = request.form['end_time']

        # Room check
        if room not in ['N410', 'N411']:
            flash("Invalid room selected.")
            return render_template('index.html', bookings=bookings)

        # Duration check
        fmt = "%H:%M"
        st = datetime.strptime(start_time, fmt)
        et = datetime.strptime(end_time, fmt)
        duration = (et - st).total_seconds() / 60
        if duration < 30:
            flash("Minimum booking duration is 30 minutes.")
            return render_template('index.html', bookings=bookings)

        # Conflict check
        if is_conflict(room, date, start_time, end_time):
            flash(f"Conflict! Room {room} is already booked during this time.")
            return render_template('index.html', bookings=bookings)

        # Save to Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([name, room, date, start_time, end_time])
        wb.save(EXCEL_FILE)

        return render_template('confirmation.html', name=name, room=room, date=date, start_time=start_time, end_time=end_time)

    return render_template('index.html', bookings=bookings)

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
